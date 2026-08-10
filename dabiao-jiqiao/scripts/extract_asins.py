# -*- coding: utf-8 -*-
"""从竞品表提取 ASIN 清单 (逗号分隔) -> asins.txt
用法: python extract_asins.py --input 竞品表.xlsx [--sheet 工作表名] [--out asins.txt] [--asin-col ASIN]
"""
import openpyxl, argparse, sys


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True, help='竞品表 xlsx 路径')
    ap.add_argument('--sheet', default=None, help='工作表名, 默认第一个 sheet')
    ap.add_argument('--out', default='asins.txt')
    ap.add_argument('--asin-col', default='ASIN')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, read_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    if args.asin_col not in header:
        print(f'[X] 表头没有列 "{args.asin_col}": {list(header)[:12]}...')
        sys.exit(1)
    i = header.index(args.asin_col)
    asins = [str(r[i]).strip() for r in rows if r[i]]
    with open(args.out, 'w', encoding='utf-8') as f:
        f.write(','.join(asins))
    print(f'[✓] {len(asins)} 个 ASIN -> {args.out}')


if __name__ == '__main__':
    main()
