# -*- coding: utf-8 -*-
"""回写: tags.json -> xlsx 指定列 (按 ASIN 匹配, 列不存在自动追加)
用法: python write_tags.py --input 竞品表.xlsx --tags tags.json --output 表-打标.xlsx
      [--sheet DE] [--asin-col ASIN] [--frame-col 相框材质] [--panel-col 面板材质]
tags.json 格式: {"ASIN": {"frame": "材质大类", "panel": "具体材质"}}
"""
import openpyxl, json, argparse, sys


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True, help='原始竞品表 xlsx')
    ap.add_argument('--tags', required=True, help='tags.json')
    ap.add_argument('--output', required=True, help='输出 xlsx 路径')
    ap.add_argument('--sheet', default=None, help='工作表名, 默认第一个 sheet')
    ap.add_argument('--asin-col', default='ASIN')
    ap.add_argument('--frame-col', default='相框材质')
    ap.add_argument('--panel-col', default='面板材质')
    args = ap.parse_args()

    tags = json.load(open(args.tags, encoding='utf-8'))
    wb = openpyxl.load_workbook(args.input)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    if args.asin_col not in header:
        print(f'[X] 表头没有 ASIN 列: {header[:12]}...'); sys.exit(1)
    # 目标列不存在则自动追加
    for col in (args.frame_col, args.panel_col):
        if col not in header:
            ws.cell(row=1, column=len(header) + 1).value = col
            header = [c.value for c in ws[1]]
    i_asin, i_f, i_p = header.index(args.asin_col), header.index(args.frame_col), header.index(args.panel_col)

    filled_f = filled_p = 0
    missing = []
    for row in ws.iter_rows(min_row=2):
        asin = row[i_asin].value
        if not asin:
            continue
        t = tags.get(str(asin))
        if t and t.get('frame'):
            row[i_f].value = t['frame']; filled_f += 1
        else:
            missing.append((asin, args.frame_col))
        if t and t.get('panel'):
            row[i_p].value = t['panel']; filled_p += 1
        else:
            missing.append((asin, args.panel_col))

    wb.save(args.output)
    total = len(tags)
    print(f'框体已填: {filled_f}/{total}  面板已填: {filled_p}/{total}')
    if missing:
        print('未填:', missing[:20], ('...' if len(missing) > 20 else ''))
    print('saved:', args.output)


if __name__ == '__main__':
    main()
