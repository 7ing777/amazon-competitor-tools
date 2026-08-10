# -*- coding: utf-8 -*-
"""
Step C: 用 data.json(+可选 content.json) 渲染竞对矩阵 xlsx (格式对齐"欧洲相框主要竞对矩阵-简")
用法:
  python render_matrix.py --data matrix_data.json [--content content.json] --out 竞对矩阵.xlsx
      [--title 德站相框矩阵] [--order "定位A,定位B,..."]
content.json 结构 (LLM 分析产出, 每定位):
  {"定位名": {"style": 风格特点, "color": 颜色系列, "scene": 使用场景, "customer": 目标客户群,
             "quality": 关键品质差异, "size_text": 主要尺寸段, "note": 竞对情况,
             "subs": ["细分类型1", "细分类型2", ...]}}
  subs 顺序对应 data.json 中该定位 brands TopN 的顺序; 缺省 content 时描述列留空、每品牌一行。
列结构: 定位|风格特点|颜色系列|使用场景|目标客户群|关键品质差异|细分类型|主要尺寸段|尺寸打勾*N|
        月销量(占比)|月销额(占比)|主要竞对|对标销量(占比)|对标销售额(占比)|代表链接(定位级合并)|竞对情况
"""
import openpyxl, json, argparse, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILLS = [PatternFill('solid', fgColor='FDE9D9'), PatternFill('solid', fgColor='DDEBF7'),
         PatternFill('solid', fgColor='E2EFDA'), PatternFill('solid', fgColor='FFF2CC'),
         PatternFill('solid', fgColor='F2DCDB'), PatternFill('solid', fgColor='E4DFEC')]

fmt_n = lambda x: f'{x:,}'


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--data', required=True, help='build_matrix_data.py 输出的 data.json')
    ap.add_argument('--content', default=None, help='LLM 分析内容 content.json(可缺省)')
    ap.add_argument('--out', required=True)
    ap.add_argument('--title', default='竞对矩阵（Top100 BSR）')
    ap.add_argument('--order', default=None, help='定位显示顺序, 逗号分隔(缺省=data.json顺序)')
    ap.add_argument('--currency', default='€', help='货币符号, 默认 €(可传 \$)')
    ap.add_argument('--market', default='欧洲', help='站点/市场名, 用于表头如: 欧洲/美国')
    args = ap.parse_args()

    data = json.load(open(args.data, encoding='utf-8'))
    content = json.load(open(args.content, encoding='utf-8')) if args.content and os.path.exists(args.content) else {}
    sizes = data.get('size_cols', [])
    if args.order:
        order = [x.strip() for x in args.order.split(',')]
        by_pos = {c['pos']: c for c in data['cats']}
        cats = [by_pos[o] for o in order if o in by_pos]
        cats += [c for c in data['cats'] if c['pos'] not in by_pos or c['pos'] not in order]
        data['cats'] = cats

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '产品定位'

    thin = Side(style='thin', color='808080')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='4472C4')
    head_font = Font(name='微软雅黑', size=9, bold=True, color='FFFFFF')
    body_font = Font(name='微软雅黑', size=9)
    bold_font = Font(name='微软雅黑', size=9, bold=True)
    wrap = Alignment(vertical='center', wrap_text=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    n_sizes = len(sizes)
    base = 8  # 定位..主要尺寸段
    n_total = base + n_sizes + 7  # + 月销量..竞对情况

    # R1 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_total)
    ws.cell(row=1, column=1, value=args.title)
    ws.cell(row=1, column=1).font = Font(name='微软雅黑', size=14, bold=True)
    ws.cell(row=1, column=1).alignment = center
    ws.row_dimensions[1].height = 24

    # R2 尺寸分组
    if n_sizes:
        groups = [(0, min(4, n_sizes)), (4, min(9, n_sizes)), (9, n_sizes)]
        labels = ['小尺寸', '中尺寸', '大尺寸']
        for gi, (c1, c2) in enumerate(groups):
            if c1 >= n_sizes:
                break
            c2 = max(c2, c1 + 1)
            ws.merge_cells(start_row=2, start_column=base + 1 + c1, end_row=2, end_column=base + 1 + c2 - 1)
            cell = ws.cell(row=2, column=base + 1 + c1, value=labels[gi])
            cell.font = head_font; cell.fill = head_fill; cell.alignment = center
        ws.row_dimensions[2].height = 18

    # R3 列头
    headers = ['定位', '风格特点', '颜色系列', '使用场景', '目标客户群', '关键品质差异', '细分类型', '主要尺寸段']
    headers += sizes
    headers += [f'{args.market}月销量（市场占比）', f'{args.market}月销售额(市场占比）', f'{args.market}市场主要竞对',
                '对标竞对销量(类目占比）', '对标竞对销售额(类目占比）', '代表链接', '竞对情况']
    r3 = 3 if n_sizes else 2
    for ci, htxt in enumerate(headers, start=1):
        cell = ws.cell(row=r3, column=ci, value=htxt)
        cell.font = head_font; cell.fill = head_fill; cell.alignment = center; cell.border = border
    ws.row_dimensions[r3].height = 40

    # 数据块
    row = r3 + 1
    for bi, c in enumerate(data['cats']):
        pos = c['pos']
        ct = content.get(pos, {})
        subs = ct.get('subs') or [None] * len(c['brands'])
        n = max(len(subs), 1)
        r0 = row
        fill = FILLS[bi % len(FILLS)]
        for si in range(n):
            r = row + si
            b = c['brands'][si] if si < len(c['brands']) else None
            sub = subs[si] if si < len(subs) else None
            ws.cell(row=r, column=1, value=pos if si == 0 else None)
            ws.cell(row=r, column=2, value=ct.get('style') if si == 0 else None)
            ws.cell(row=r, column=3, value=ct.get('color') if si == 0 else None)
            ws.cell(row=r, column=4, value=ct.get('scene') if si == 0 else None)
            ws.cell(row=r, column=5, value=ct.get('customer') if si == 0 else None)
            ws.cell(row=r, column=6, value=ct.get('quality') if si == 0 else None)
            ws.cell(row=r, column=7, value=sub)
            ws.cell(row=r, column=8, value=ct.get('size_text') if si == 0 else None)
            for ci in range(base + 1, base + 1 + n_sizes):
                name = sizes[ci - base - 1]
                ws.cell(row=r, column=ci, value='√' if si == 0 and name in c.get('size_marks', []) else None)
            ws.cell(row=r, column=base + n_sizes + 1, value=f"{fmt_n(c['sales'])}（{c['sales_share']}%）" if si == 0 else None)
            ws.cell(row=r, column=base + n_sizes + 2, value=f"{fmt_n(c['rev'])}{args.currency}（{c['rev_share']}%）" if si == 0 else None)
            ws.cell(row=r, column=base + n_sizes + 3, value=b['brand'] if b else None)
            ws.cell(row=r, column=base + n_sizes + 4, value=f"{fmt_n(b['sales'])}（{b['sales_share']}%）" if b else None)
            ws.cell(row=r, column=base + n_sizes + 5, value=f"{fmt_n(b['rev'])}（{b['rev_share']}%）" if b else None)
            ws.cell(row=r, column=base + n_sizes + 6, value=c['rep_link'] if si == 0 else None)
            ws.cell(row=r, column=base + n_sizes + 7, value=ct.get('note') if si == 0 else None)
            for ci in range(1, n_total + 1):
                cell = ws.cell(row=r, column=ci)
                cell.border = border
                cell.font = bold_font if ci == 1 else body_font
                cell.alignment = center if base < ci <= base + n_sizes else wrap
                if ci in (1, 2, 3, 4, 5, 6, 8, base + n_sizes + 1, base + n_sizes + 2, base + n_sizes + 7):
                    cell.fill = fill
        for ci in (1, 2, 3, 4, 5, 6, 8, base + n_sizes + 1, base + n_sizes + 2, base + n_sizes + 6, base + n_sizes + 7):
            if n > 1:
                ws.merge_cells(start_row=r0, start_column=ci, end_row=r0 + n - 1, end_column=ci)
        row += n

    # 列宽
    widths = {1: 10, 2: 18, 3: 20, 4: 18, 5: 16, 6: 20, 7: 22, 8: 16}
    for ci in range(base + 1, base + 1 + n_sizes):
        widths[ci] = 8.5
    widths.update({base + n_sizes + 1: 15, base + n_sizes + 2: 17, base + n_sizes + 3: 12,
                   base + n_sizes + 4: 16, base + n_sizes + 5: 18, base + n_sizes + 6: 30, base + n_sizes + 7: 34})
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    for r in range(r3 + 1, row):
        ws.row_dimensions[r].height = 46
    ws.freeze_panes = 'D%d' % (r3 + 1)

    wb.save(args.out)
    print(f'[✓] saved: {args.out}  ({len(data["cats"])} 定位, {n_total} 列)')


if __name__ == '__main__':
    main()
