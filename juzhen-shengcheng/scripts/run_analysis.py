# -*- coding: utf-8 -*-
"""
一键分析入口 (给其他工作人员用, 无需 Hermes/亚马逊账号/浏览器)
用法:
  python run_analysis.py --input 打标表.xlsx [--config config.ini] [--out-prefix 输出名前缀]
流程(自动):
  1. 读取打标表, 校验列名(可改 config.ini)
  2. 生成市场数据 data.json (build_matrix_data)
  3. 分类分析: 优先 透视表版(make_pivot, 需本机Excel/WPS+pywin32); 失败自动降级 静态版(analyze_categories)
  4. 自动生成 content.json (规则总结, 无需LLM)
  5. 渲染 竞对矩阵 (render_matrix)
输出: <前缀>-透视表版.xlsx 或 <前缀>-市场分析.xlsx + <前缀>竞对矩阵.xlsx
"""
import argparse, configparser, json, os, subprocess, sys, re
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
PY = sys.executable


def load_cfg(path):
    cfg = configparser.ConfigParser()
    cfg.read(path, encoding='utf-8')
    c = dict(cfg.items('columns')) if cfg.has_section('columns') else {}
    p = dict(cfg.items('params')) if cfg.has_section('params') else {}
    return c, p


def header_of(path, sheet):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    return [str(x) for x in rows[0]], ws.max_row, ws.max_column, wb.sheetnames


def auto_detect(header, cfg_cols, rows=None):
    """列名: 配置优先, 否则常见别名猜测; size_col 仅当值为 cm 尺寸格式才启用"""
    aliases = {
        'cat_col': ['打开', '相框类型', '分类', '定位', '类别'],
        'brand_col': ['品牌', 'Brand'],
        'asin_col': ['ASIN'],
        'sales_col': ['月销量'],
        'rev_col': ['月销售额($)', '月销售额(€)', '月销售额', '月销额'],
        'size_col': ['尺寸', '商品尺寸'],
        'price_col': ['价格($)', '价格(€)', '价格'],
        'dim2_col': ['设计结构', '材质', '细分'],
    }
    out = {}
    for key, al in aliases.items():
        cfg_v = (cfg_cols.get(key) or '').strip()
        # config 列名仅当真实存在于表头才生效, 否则回退别名自动识别
        out[key] = cfg_v if cfg_v and cfg_v in header else next((a for a in al if a in header), '')
    # size_col 校验: 采样该列值, 必须含 cm 尺寸格式(如 50x70 / 21×29.7), 否则关闭
    if out['size_col'] and rows:
        si = header.index(out['size_col'])
        has_cm = any(re.search(r'\d+\s*[x×*]\s*\d+', str(r[si])) for r in rows[1:30] if r[si] is not None)
        if not has_cm:
            out['size_col'] = ''
    return out


def make_content(data_path, order):
    """自动生成 content.json: subs=品牌Top, note=规则总结 (无需LLM)"""
    d = json.load(open(data_path, encoding='utf-8'))
    cats = sorted(d['cats'], key=lambda x: order.index(x['pos']) if order and x['pos'] in order else 99)
    content = {}
    for c in cats:
        items = c['brands']
        if not items:
            continue
        top1, top3 = items[0], items[:3]
        cr1_v = top1['rev_share']
        pattern = '显著寡头垄断' if cr1_v >= 40 else ('头部集中' if cr1_v >= 20 else '格局分散')
        tactic = ('高客单' if top1['rev_share'] > top1['sales_share'] * 1.05 else
                  '走量低价' if top1['sales_share'] > top1['rev_share'] * 1.05 else '量额均衡')
        note = f'{c["pos"]}市场{pattern}：头部 {top1["brand"]} 占 {cr1_v}% 销售额（{tactic}），前3合计约 {sum(b["rev_share"] for b in top3)}% 销售额。'
        content[c['pos']] = {
            'style': '', 'color': '', 'scene': '', 'customer': '', 'quality': '',
            'size_text': '', 'note': note,
            'subs': [f'{b["brand"]}（头部链接）' for b in items[:3]],
        }
    return content


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True)
    ap.add_argument('--config', default=os.path.join(HERE, 'config.ini'))
    ap.add_argument('--out-prefix', default=None)
    ap.add_argument('--sheet', default=None)
    ap.add_argument('--skip-pivot', action='store_true')
    args = ap.parse_args()

    cfg_cols, cfg_params = load_cfg(args.config)
    out_prefix = args.out_prefix or os.path.splitext(args.input)[0]
    header, max_row, max_col, sheets = header_of(args.input, args.sheet)
    sheet_arg = args.sheet or sheets[0]
    # 读行(定位识别 + 列名采样)
    import openpyxl
    wb = openpyxl.load_workbook(args.input, read_only=True)
    ws = wb[sheet_arg]
    rows = list(ws.iter_rows(values_only=True))
    h = rows[0]
    cols = auto_detect(header, cfg_cols, rows=rows)
    missing = [k for k, v in cols.items() if not v and k not in ('size_col', 'price_col', 'dim2_col')]
    if missing:
        print('[X] 缺少列: %s\n    表头: %s' % (', '.join(missing), header[:20]))
        sys.exit(1)

    # 定位列表自动识别 + 实际数据行数(按 ASIN 非空, 不假设最后一行是说明行)
    ci = h.index(cols['cat_col'])
    cats = []
    data_rows = 0
    asin_i = h.index(cols['asin_col']) if cols['asin_col'] in h else None
    for r in rows[1:]:
        if asin_i is not None and r[asin_i] is None:
            continue
        if not any(v is not None for v in r):
            continue
        data_rows += 1
        if r[ci]:
            v = str(r[ci]).strip()
            if v not in cats:
                cats.append(v)
    wb.close()
    data_rows = max_row - 1  # 最后一行多为说明行
    currency = cfg_params.get('currency', '€')
    market = cfg_params.get('market', '欧洲')
    title = cfg_params.get('title', f'{market}竞对矩阵（TOP{data_rows}）')
    suffix = cfg_params.get('sheet_suffix', '')
    order = cfg_params.get('order', '') or ','.join(cats)

    print(f'[1/5] 数据: {len(cats)} 个定位 {", ".join(cats)} | 列: {cols}')
    # Step A
    link_args = ['--link-col', '商品详情页链接'] if '商品详情页链接' in header else []
    size_args = ['--size-col', cols['size_col']] if cols['size_col'] else ['--no-sizes']
    subprocess.run([PY, os.path.join(HERE, 'build_matrix_data.py'), '--input', args.input,
                    '--sheet', sheet_arg, '--cat-col', cols['cat_col'], '--brand-col', cols['brand_col'],
                    '--sales-col', cols['sales_col'], '--rev-col', cols['rev_col'], '--asin-col', cols['asin_col'],
                    '--out', f'{out_prefix}_data.json'] + link_args + size_args,
                   check=False)
    # Step A2p: 透视表版
    pivot_ok = False
    if not args.skip_pivot:
        r = subprocess.run([PY, os.path.join(HERE, 'make_pivot.py'), '--input', os.path.abspath(args.input),
                            '--out', os.path.abspath(f'{out_prefix}-透视表版.xlsx'), '--source-sheet', sheet_arg,
                            '--data-rows', str(data_rows), '--data-cols', str(max_col),
                            '--cat-col', cols['cat_col'], '--brand-col', cols['brand_col'], '--asin-col', cols['asin_col'],
                            '--sales-col', cols['sales_col'], '--rev-col', cols['rev_col'],
                            '--cats', ','.join(cats), '--sheet-suffix', suffix,
                            '--dim2-col', cols['dim2_col'] or '',
                            '--dim2-sheet', cfg_params.get('dim2_sheet', '设计结构占比分析'),
                            '--share-sheet', cfg_params.get('share_sheet', '分类占比')],
                           capture_output=True, text=True)
        if r.returncode == 0 and os.path.exists(f'{out_prefix}-透视表版.xlsx'):
            pivot_ok = True
            print(f'[2/5] 透视表版: ✓ {os.path.basename(out_prefix)}-透视表版.xlsx')
        else:
            print('[2/5] 透视表版失败(本机无Excel/WPS或pywin32?):', (r.stderr or r.stdout)[-160:])
    if not pivot_ok:
        subprocess.run([PY, os.path.join(HERE, 'analyze_categories.py'), '--input', args.input,
                        '--sheet', sheet_arg, '--cat-col', cols['cat_col'], '--brand-col', cols['brand_col'],
                        '--sales-col', cols['sales_col'], '--rev-col', cols['rev_col'], '--asin-col', cols['asin_col'],
                        '--dim2-col', cols['dim2_col'] or '', '--price-col', cols['price_col'] or '',
                        '--out', f'{out_prefix}-市场分析.xlsx'], check=False)
        print(f'[2/5] 已降级为静态版: {os.path.basename(out_prefix)}-市场分析.xlsx')
    # Step B: 自动 content.json
    content = make_content(f'{out_prefix}_data.json', order.split(','))
    with open(f'{out_prefix}_content.json', 'w', encoding='utf-8') as f:
        json.dump(content, f, ensure_ascii=False, indent=1)
    print('[3/5] content.json 已自动生成(规则总结)')
    # Step C: 矩阵
    subprocess.run([PY, os.path.join(HERE, 'render_matrix.py'), '--data', f'{out_prefix}_data.json',
                    '--content', f'{out_prefix}_content.json', '--out', f'{out_prefix}竞对矩阵.xlsx',
                    '--title', title, '--currency', currency, '--market', market, '--order', order],
                   check=False)
    print(f'[4/5] 竞对矩阵: ✓ {os.path.basename(out_prefix)}竞对矩阵.xlsx')
    print('[5/5] 完成。交付文件:')
    print(f'  - {os.path.basename(out_prefix)}-透视表版.xlsx / -市场分析.xlsx (分类分析)')
    print(f'  - {os.path.basename(out_prefix)}竞对矩阵.xlsx (竞对矩阵)')


if __name__ == '__main__':
    main()
