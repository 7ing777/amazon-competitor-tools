# -*- coding: utf-8 -*-
"""
Step A2: 打标表 → 市场分析 sheets (格式精确对齐"塑料垃圾桶 细分类目"示例, 追加到打标表副本)
追加生成 sheets:
  <定位>垃圾桶 × N  — 垄断性分析: 行标签|求和项:月销售额($)|求和项:月销量2|销售额占比|销量占比|总结
  设计结构占比分析    — 定位→结构→ASIN 层级, 占比=占类目全量
  销量占比对比图      — 折线图: 各定位 销量占比 vs 销售额占比
  价格分析           — 散点图: 价格($) × 月销量, 按定位分组着色
用法:
  python analyze_categories.py --input 打标表.xlsx [--sheet 工作表] [--out 输出.xlsx]
      --cat-col 打开 --brand-col 品牌 --sales-col 月销量 --rev-col "月销售额($)"
      [--asin-col ASIN] [--dim2-col 设计结构] [--price-col "价格($)"]
说明: 输出 = 原表 + 追加分析 sheets (用户口径: 直接加到打标表上即可);
      占比为小数原值(0.1093)配百分比格式, 与用户示例一致。
"""
import openpyxl, argparse, sys, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THIN = Side(style='thin', color='808080')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill('solid', fgColor='4472C4')
HEAD_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
BODY_FONT = Font(name='微软雅黑', size=10)
BOLD_FONT = Font(name='微软雅黑', size=10, bold=True)
WRAP = Alignment(vertical='center', wrap_text=True)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
PCT = '0.00%'
NUM = '#,##0'
NUM2 = '#,##0.00'


def style_sheet(ws, widths, last_col_letter):
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def make_summary(pos, n, sales, rev, items):
    """规则化生成垄断性分析总结 (ASIN级)"""
    if not items:
        return f'{pos}：该定位暂无明细数据。'
    tot_s = sum(i['sales'] for i in items)
    tot_v = sum(i['rev'] for i in items)
    top1, top3 = items[0], items[:3]
    cr1_s = top1['sales'] / tot_s * 100
    cr1_v = top1['rev'] / tot_v * 100
    cr3_v = sum(i['rev'] for i in top3) / tot_v * 100
    if cr1_v >= 40:
        pattern = '显著寡头垄断格局'
    elif cr1_v >= 20:
        pattern = '头部集中特征明显'
    else:
        pattern = '格局相对分散'
    tactic = ''
    if top1['sales'] / tot_s > top1['rev'] / tot_v * 1.05:
        tactic = f'{top1["brand"]} 走量优势突出但客单价偏低'
    elif top1['rev'] / tot_v > top1['sales'] / tot_s * 1.05:
        tactic = f'{top1["brand"]} 销量占比不及销售额占比，依靠更高客单价拉动'
    else:
        tactic = f'{top1["brand"]} 销量与销售额占比均衡'
    tail = ''
    if len(items) > 3:
        tail = f'；其余 {len(items) - 3} 条链接占比多在 1%-{max(2, int(items[3]["rev"] / tot_v * 100))}% 区间，长尾分散'
    return (f'{pos}市场呈现{pattern}：头部单品 {top1["asin"]} 独占 {cr1_v:.0f}% 销售额、{cr1_s:.0f}% 销量，'
            f'前三款合计瓜分 {cr3_v:.0f}% 销售额{tail}；{tactic}。')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True)
    ap.add_argument('--sheet', default=None)
    ap.add_argument('--out', default=None, help='缺省 = 输入文件名-市场分析.xlsx')
    ap.add_argument('--cat-col', default='打开')
    ap.add_argument('--brand-col', default='品牌')
    ap.add_argument('--sales-col', default='月销量')
    ap.add_argument('--rev-col', default='月销售额($)')
    ap.add_argument('--asin-col', default='ASIN')
    ap.add_argument('--dim2-col', default='设计结构', help='次级维度列, None 关闭')
    ap.add_argument('--price-col', default='价格($)', help='价格列(散点图用)')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input)
    src_ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    rows = list(src_ws.iter_rows(values_only=True))
    header = [str(x) for x in rows[0]]
    for col in (args.cat_col, args.brand_col, args.sales_col, args.rev_col):
        if col not in header:
            print(f'[X] 表头缺少列 "{col}"。可用列: {header[:15]}...')
            sys.exit(1)
    gi = {col: header.index(col) for col in (args.cat_col, args.brand_col, args.sales_col, args.rev_col, args.asin_col)}
    for col, key in ((args.dim2_col, 'd2'), (args.price_col, 'price')):
        if col and col in header:
            gi[key] = header.index(col)

    cats = {}
    grand = {'sales': 0.0, 'rev': 0.0, 'n': 0}
    for r in rows[1:]:
        if not r[gi[args.asin_col]]:
            continue
        pos = str(r[gi[args.cat_col]]).strip() if r[gi[args.cat_col]] else '(未分类)'
        sales = float(r[gi[args.sales_col]] or 0)
        rev = float(r[gi[args.rev_col]] or 0)
        brand = str(r[gi[args.brand_col]] or '').strip()
        asin = str(r[gi[args.asin_col]]).strip()
        d2 = str(r[gi['d2']]).strip() if 'd2' in gi and r[gi['d2']] else '(未标注)'
        price = float(r[gi['price']]) if 'price' in gi and r[gi['price']] is not None else None
        c = cats.setdefault(pos, {'n': 0, 'sales': 0.0, 'rev': 0.0, 'items': [], 'd2': {}})
        c['n'] += 1; c['sales'] += sales; c['rev'] += rev
        c['items'].append({'asin': asin, 'brand': brand, 'sales': sales, 'rev': rev, 'price': price})
        d2c = c['d2'].setdefault(d2, {'n': 0, 'sales': 0.0, 'rev': 0.0, 'items': []})
        d2c['n'] += 1; d2c['sales'] += sales; d2c['rev'] += rev
        d2c['items'].append({'asin': asin, 'brand': brand, 'sales': sales, 'rev': rev})
        grand['sales'] += sales; grand['rev'] += rev; grand['n'] += 1
    for c in cats.values():
        c['items'].sort(key=lambda x: -x['rev'])
    cats = dict(sorted(cats.items(), key=lambda x: -x[1]['rev']))

    # 每定位 垄断性分析 (格式对齐示例)
    for pos, c in cats.items():
        name = f'{pos}垃圾桶'
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws.merge_cells('A1:E1')
        ws['A1'] = '垄断性分析'
        ws['A1'].font = Font(name='微软雅黑', size=12, bold=True)
        ws.merge_cells('F3:F4')
        ws['F3'] = f'总结：{make_summary(pos, c["n"], c["sales"], c["rev"], c["items"])}'
        ws['F3'].font = BODY_FONT
        ws['F3'].alignment = WRAP
        hdr = ['行标签', '求和项:月销售额($)', '求和项:月销量2', '销售额占比', '销量占比']
        for ci, h in enumerate(hdr, start=1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = HEAD_FONT; cell.fill = HEAD_FILL; cell.alignment = CENTER; cell.border = BORDER
        ws.cell(row=4, column=1, value=pos).font = BOLD_FONT
        ws.cell(row=4, column=2, value=round(c['rev'], 2)).number_format = NUM2
        ws.cell(row=4, column=3, value=round(c['sales'])).number_format = NUM
        for ci in range(1, 6):
            ws.cell(row=4, column=ci).border = BORDER
        for it in c['items']:
            ws.append([it['asin'], round(it['rev'], 2), round(it['sales']),
                       it['rev'] / c['rev'], it['sales'] / c['sales']])
            r = ws.max_row
            for ci in range(1, 6):
                cell = ws.cell(row=r, column=ci)
                cell.border = BORDER
                if ci == 2:
                    cell.number_format = NUM2
                elif ci == 3:
                    cell.number_format = NUM
                elif ci in (4, 5):
                    cell.number_format = PCT
                    cell.alignment = CENTER
        style_sheet(ws, [16, 20, 14, 13, 13, 60], 'F')
        ws.column_dimensions['F'].width = 70
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 20
        ws.freeze_panes = 'A5'

    # 设计结构占比分析
    if args.dim2_col:
        name = f'{args.dim2_col}占比分析'
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws.merge_cells('A1:E1')
        ws['A1'] = f'{args.dim2_col}占比分析'
        ws['A1'].font = Font(name='微软雅黑', size=12, bold=True)
        ws.merge_cells('F2:F3')
        ws['F2'] = '总结：按结构拆分各定位销售额构成；占比基数为表内全部产品合计。'
        ws['F2'].font = BODY_FONT
        ws['F2'].alignment = WRAP
        hdr = ['行标签', '求和项:月销量2', '求和项:月销售额($)', '销量占比', '销售额占比']
        for ci, h in enumerate(hdr, start=1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = HEAD_FONT; cell.fill = HEAD_FILL; cell.alignment = CENTER; cell.border = BORDER
        r = 3
        for pos, c in cats.items():
            ws.cell(row=r, column=1, value=pos).font = BOLD_FONT
            ws.cell(row=r, column=2, value=round(c['sales'])).number_format = NUM
            ws.cell(row=r, column=3, value=round(c['rev'], 2)).number_format = NUM2
            for ci in range(1, 6):
                ws.cell(row=r, column=ci).border = BORDER
            r += 1
            for d2, d2c in sorted(c['d2'].items(), key=lambda x: -x[1]['rev']):
                ws.cell(row=r, column=1, value=d2).font = BODY_FONT
                ws.cell(row=r, column=2, value=round(d2c['sales'])).number_format = NUM
                ws.cell(row=r, column=3, value=round(d2c['rev'], 2)).number_format = NUM2
                ws.cell(row=r, column=4, value=d2c['sales'] / grand['sales']).number_format = PCT
                ws.cell(row=r, column=5, value=d2c['rev'] / grand['rev']).number_format = PCT
                for ci in range(1, 6):
                    ws.cell(row=r, column=ci).border = BORDER
                r += 1
                for it in d2c['items']:
                    ws.cell(row=r, column=1, value=it['asin']).font = BODY_FONT
                    ws.cell(row=r, column=2, value=round(it['sales'])).number_format = NUM
                    ws.cell(row=r, column=3, value=round(it['rev'], 2)).number_format = NUM2
                    ws.cell(row=r, column=4, value=it['sales'] / grand['sales']).number_format = PCT
                    ws.cell(row=r, column=5, value=it['rev'] / grand['rev']).number_format = PCT
                    for ci in range(1, 6):
                        ws.cell(row=r, column=ci).border = BORDER
                    r += 1
        style_sheet(ws, [16, 14, 20, 13, 13, 60], 'F')
        ws.column_dimensions['F'].width = 70
        ws.freeze_panes = 'A3'

    # 销量占比对比图 (折线图)
    from openpyxl.chart import LineChart, Reference
    name = '销量占比对比图'
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(['定位', '销量占比', '销售额占比'])
    for pos, c in cats.items():
        ws.append([pos, c['sales'] / grand['sales'], c['rev'] / grand['rev']])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
            if cell.row == 1:
                cell.font = HEAD_FONT; cell.fill = HEAD_FILL
            if cell.column in (2, 3) and cell.row > 1:
                cell.number_format = PCT
    chart = LineChart()
    chart.title = '各定位销量占比 vs 销售额占比'
    chart.y_axis.numFmt = '0%'
    chart.y_axis.majorGridlines = None
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 10
    chart.width = 24
    ws.add_chart(chart, 'E2')

    # 价格分析 (散点图, 按定位分组)
    from openpyxl.chart import ScatterChart, Reference, Series
    name = '价格分析'
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    colors = ['FF0000', '0070C0', '00B050', 'FFC000', '7030A0', 'FF6600', '808080']
    for ci, (pos, c) in enumerate(cats.items()):
        col_x = 2 + ci * 3
        col_y = col_x + 1
        ws.cell(row=1, column=col_x, value=f'{pos} 价格')
        ws.cell(row=1, column=col_y, value=f'{pos} 月销量')
        row = 2
        for it in c['items']:
            if it['price'] is None or it['price'] <= 0:
                continue
            ws.cell(row=row, column=col_x, value=it['price']).number_format = '#,##0.00'
            ws.cell(row=row, column=col_y, value=it['sales']).number_format = NUM
            row += 1
    chart = ScatterChart()
    chart.title = '价格($) × 月销量（按定位分组）'
    chart.x_axis.title = '价格($)'
    chart.y_axis.title = '月销量'
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.style = 13
    from openpyxl.chart.marker import Marker
    for ci, (pos, c) in enumerate(cats.items()):
        col_x = 2 + ci * 3
        col_y = col_x + 1
        nrows = 1 + sum(1 for it in c['items'] if it['price'] and it['price'] > 0)
        if nrows < 2:
            continue
        xref = Reference(ws, min_col=col_x, min_row=2, max_row=nrows)
        yref = Reference(ws, min_col=col_y, min_row=2, max_row=nrows)
        s = Series(yref, xref, title=pos)
        s.marker = Marker(symbol='circle', size=6)
        s.graphicalProperties.solidFill = colors[ci % len(colors)]
        chart.series.append(s)
    chart.height = 11
    chart.width = 26
    ws.add_chart(chart, 'J2')

    out = args.out or (os.path.splitext(args.input)[0] + '-市场分析.xlsx')
    wb.save(out)
    print(f'[√] saved: {out}  (追加 {len(cats) + 3} 个 sheets 到原表)')
    print(f'  垄断性分析 ×{len(cats)}: {", ".join(cats.keys())}')
    print(f'  {args.dim2_col}占比分析 | 销量占比对比图 | 价格分析')


if __name__ == '__main__':
    main()
