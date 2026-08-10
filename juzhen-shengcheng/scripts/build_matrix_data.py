# -*- coding: utf-8 -*-
"""
Step A: 从任意"已打标竞品表"自动计算市场数据 -> data.json (通用脚本, 换类目零改动)
用法:
  python build_matrix_data.py --input 竞品表.xlsx [--sheet 工作表] [--out matrix_data.json]
      [--cat-col 相框类型] [--brand-col 品牌] [--sales-col 月销量] [--rev-col 月销售额(€)]
      [--size-col 商品尺寸] [--link-col 商品详情页链接] [--no-sizes] [--sizes-file 自定义尺寸.json]
输出 matrix_data.json:
  grand(总量) + cats[]: 每定位 n/销量/销额/占比/品牌Top(含类目内占比)/代表链接(定位内销额最高ASIN)/尺寸打勾
注意: 占比直接由表内数据自算, 与卖家精灵分析表口径一致(已用相框100验证: 21/52/23/4%)
"""
import openpyxl, json, argparse, re, sys


def parse_cm(size_str):
    m = re.findall(r'(\d+[.,]?\d*)', str(size_str))
    if len(m) < 2:
        return None
    a, b = float(m[0].replace(',', '.')), float(m[1].replace(',', '.'))
    if a < 8 or b < 8 or a > 150 or b > 150:
        return None  # 脏数据(如 40x1.2)
    return (min(a, b), max(a, b))


def load_sizes(path=None):
    default = [
        ('4*6', (10.2, 15.2)), ('5*7', (12.7, 17.8)), ('8*10', (20.3, 25.4)), ('8*8', (20.3, 20.3)),
        ('10*10', (25.4, 25.4)), ('A4-8.3*11.7(欧)', (21.0, 29.7)), ('11*17', (27.9, 43.2)), ('A3-11.7*16.5(欧)', (29.7, 42.0)),
        ('12*12(欧)', (30.5, 30.5)), ('12*18', (30.5, 45.7)), ('16*20', (40.6, 50.8)), ('16*24', (40.6, 61.0)),
        ('A2-16.5*23.4(欧)', (42.0, 59.4)), ('18*24', (45.7, 61.0)), ('20*28', (50.8, 71.1)), ('20*30', (50.8, 76.2)),
        ('24*36', (61.0, 91.4)),
    ]
    if path:
        return [(k, tuple(v)) for k, v in json.load(open(path, encoding='utf-8')).items()]
    return default


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True)
    ap.add_argument('--sheet', default=None, help='工作表名, 默认第一个')
    ap.add_argument('--out', default='matrix_data.json')
    ap.add_argument('--cat-col', default='相框类型')
    ap.add_argument('--brand-col', default='品牌')
    ap.add_argument('--sales-col', default='月销量')
    ap.add_argument('--rev-col', default='月销售额(€)')
    ap.add_argument('--size-col', default='商品尺寸')
    ap.add_argument('--link-col', default=None, help='详情页链接列(缺省则用 dp/ASIN 拼接)')
    ap.add_argument('--asin-col', default='ASIN', help='ASIN 列(用于拼接代表链接)')
    ap.add_argument('--no-sizes', action='store_true', help='不生成尺寸打勾列')
    ap.add_argument('--sizes-file', default=None)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, read_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(x) for x in rows[0]]
    for col in (args.cat_col, args.brand_col, args.sales_col, args.rev_col):
        if col not in header:
            print(f'[X] 表头缺少列 "{col}"。可用列: {header[:15]}...')
            sys.exit(1)
    gi = {col: header.index(col) for col in (args.cat_col, args.brand_col, args.sales_col, args.rev_col)}
    if args.asin_col in header:
        gi[args.asin_col] = header.index(args.asin_col)
    if args.size_col and args.size_col in header:
        gi[args.size_col] = header.index(args.size_col)
    if args.link_col and args.link_col in header:
        gi[args.link_col] = header.index(args.link_col)

    sizes = load_sizes(args.sizes_file) if not args.no_sizes else []

    cats = {}
    grand = {'sales': 0, 'rev': 0, 'n': 0}
    for r in rows[1:]:
        if not r[gi[args.cat_col]] or not r[gi[args.brand_col]]:
            continue
        pos = str(r[gi[args.cat_col]]).strip()
        brand = str(r[gi[args.brand_col]]).strip()
        sales = float(r[gi[args.sales_col]] or 0)
        rev = float(r[gi[args.rev_col]] or 0)
        asin = str(r[gi[args.asin_col]]) if args.asin_col in header else ''
        link = ''
        if args.link_col and args.link_col in header:
            link = str(r[gi[args.link_col]] or '')
        if not asin:
            m = re.search(r'/(?:dp|gp/product)/([A-Z0-9]{10})', link)
            asin = m.group(1) if m else ''
        if not link and asin:
            link = f'https://www.amazon.de/dp/{asin}'
        d = cats.setdefault(pos, {'n': 0, 'sales': 0, 'rev': 0, 'brands': {}, 'top_asin': None, 'top_rev': 0, 'rep_link': '', 'size_marks': set()})
        d['n'] += 1
        d['sales'] += sales
        d['rev'] += rev
        b = d['brands'].setdefault(brand, {'sales': 0, 'rev': 0, 'top_asin': '', 'top_rev': 0, 'rep_link': ''})
        b['sales'] += sales
        b['rev'] += rev
        if rev > b['top_rev']:
            b['top_rev'] = rev
            b['top_asin'] = asin
            b['rep_link'] = link
        if rev > d['top_rev']:
            d['top_rev'] = rev
            d['top_asin'] = asin
            d['rep_link'] = link
        if args.size_col and args.size_col in header and sizes:
            dims = parse_cm(r[gi[args.size_col]])
            if dims:
                best, bd = None, 1e9
                for name, (sw, sh) in sizes:
                    dd = abs(dims[0] - sw) / sw + abs(dims[1] - sh) / sh
                    if dd < bd:
                        bd, best = dd, name
                if bd < 0.35:
                    d['size_marks'].add(best)
        grand['sales'] += sales
        grand['rev'] += rev
        grand['n'] += 1

    out_cats = []
    for pos, d in cats.items():
        brands = sorted(d['brands'].items(), key=lambda x: -x[1]['rev'])
        out_cats.append({
            'pos': pos, 'n': d['n'],
            'sales': round(d['sales']), 'rev': round(d['rev']),
            'sales_share': round(d['sales'] / grand['sales'] * 100),
            'rev_share': round(d['rev'] / grand['rev'] * 100),
            'brands': [{'brand': b, 'sales': round(v['sales']), 'rev': round(v['rev']),
                        'sales_share': round(v['sales'] / d['sales'] * 100),
                        'rev_share': round(v['rev'] / d['rev'] * 100),
                        'rep_asin': v['top_asin'], 'rep_link': v['rep_link']} for b, v in brands],
            'rep_asin': d['top_asin'], 'rep_link': d['rep_link'],
            'size_marks': sorted(d['size_marks']),
        })
    out_cats.sort(key=lambda x: -x['rev'])

    data = {
        'grand': {'sales': round(grand['sales']), 'rev': round(grand['rev']), 'n': grand['n']},
        'size_cols': [s[0] for s in sizes] if sizes else [],
        'cats': out_cats,
    }
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    print(f'[✓] {grand["n"]} 产品 / {len(out_cats)} 定位 -> {args.out}')
    for c in out_cats:
        print(f"  {c['pos']}: {c['n']}个 销{c['sales']:,}({c['sales_share']}%) 额{c['rev']:,}€({c['rev_share']}%) | 代表 {c['rep_asin']} | Top: {', '.join(b['brand'] for b in c['brands'][:3])}")


if __name__ == '__main__':
    main()
