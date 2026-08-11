# -*- coding: utf-8 -*-
"""
Step A3: 把分析 sheets 的静态数值改写为实时公式 (SUMIFS 引用源表, 效果等同透视表)
用法:
  python formula_analysis.py --input 分析工作簿.xlsx --source-sheet Product-DE-Last-30-days \
      --cat-col 相框类型 --brand-col 品牌 --asin-col ASIN --sales-col 月销量 --rev-col "月销售额(€)" \
      --data-start 2 --data-end 71 \
      --sheets "高端实木相框市场概况,中高端相框市场概况,窄边相框市场情况,特定功能相框市场情况,分类占比"
说明:
  - 自动识别: 定位合计行(colA∈源表分类值) / ASIN明细行 / 品牌明细行(非分类非ASIN)
  - 定位合计行→SUMIFS(单条件); 品牌行→SUMIFS(定位+品牌双条件); ASIN行→SUMIFS(单条件)
  - 占比基数自动判定: within(类目内=引用合计行) vs grand(全量=SUM源表), 双遍扫描
  - 文本/总结单元格保留不动; 输出 = 输入-公式版.xlsx
  - 注意: sheet名带尾随空格时按 strip 匹配; 打开Excel后自动重算(openpyxl写入公式无缓存值)
"""
import openpyxl, argparse, re, sys, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ASIN_RE = re.compile(r'^[A-Z0-9]{10}$')
THIN = Side(style='thin', color='808080')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = '0.00%'
NUM = '#,##0'
NUM2 = '#,##0.00'


def col_letter(header, name):
    if name not in header:
        return None
    return get_column_letter(header.index(name) + 1)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True)
    ap.add_argument('--source-sheet', required=True, help='源数据表(打标表)')
    ap.add_argument('--cat-col', required=True)
    ap.add_argument('--brand-col', default='品牌')
    ap.add_argument('--asin-col', default='ASIN')
    ap.add_argument('--sales-col', required=True)
    ap.add_argument('--rev-col', required=True)
    ap.add_argument('--data-start', type=int, required=True, help='源表数据起始行(含)')
    ap.add_argument('--data-end', type=int, required=True, help='源表数据结束行(含)')
    ap.add_argument('--sheets', required=True, help='要转换的分析sheet名, 逗号分隔')
    ap.add_argument('--out', default=None)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input)
    src = wb[args.source_sheet]
    src_rows = list(src.iter_rows(values_only=True))
    src_header = [str(x) for x in src_rows[0]]
    cl = {name: col_letter(src_header, name) for name in
          (args.cat_col, args.brand_col, args.asin_col, args.sales_col, args.rev_col)}
    for name, letter in cl.items():
        if not letter:
            print(f'[X] 源表缺少列 "{name}"')
            sys.exit(1)
    L = args.data_start
    R = args.data_end
    s = f"'{args.source_sheet}'"
    cat_set = set()
    grand_s = grand_r = 0.0
    for row in src_rows[L - 1:R]:
        cat = row[src_header.index(args.cat_col)]
        if cat:
            cat_set.add(str(cat).strip())
        try:
            grand_s += float(row[src_header.index(args.sales_col)] or 0)
        except (TypeError, ValueError):
            pass
        try:
            grand_r += float(row[src_header.index(args.rev_col)] or 0)
        except (TypeError, ValueError):
            pass

    def sumif(letter, *crit):
        parts = [f'{s}!${letter}${L}:${letter}${R}']
        parts.append(f'{s}!${cl[args.cat_col]}${L}:${cl[args.cat_col]}${R}, {crit[0]}')
        if len(crit) > 1:
            parts.append(f'{s}!${cl[args.brand_col]}${L}:${cl[args.brand_col]}${R}, {crit[1]}')
        return '=SUMIFS(' + ', '.join(parts) + ')'

    converted = 0
    targets = [x.strip() for x in args.sheets.split(',')]
    for sn in wb.sheetnames:
        if sn.strip() not in targets:
            continue
        ws = wb[sn]
        # Pass 1: 定位合计行 + 判定占比基数模式 (within=类目内 / grand=全量)
        cur_total = None
        base_mode = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a is None or str(a).strip() in ('', '行标签'):
                continue
            label = str(a).strip()
            b = ws.cell(row=r, column=2).value
            d_old = ws.cell(row=r, column=4).value
            is_num = isinstance(b, (int, float)) or (isinstance(b, str) and b.replace('.', '').replace(',', '').isdigit())
            if not is_num:
                continue
            if label in cat_set:
                cur_total = (r, label, b)
            elif base_mode is None and isinstance(d_old, (int, float)) and d_old > 0 and float(b) > 0:
                base = float(b) / d_old
                base_mode = 'within' if (cur_total and cur_total[2] and abs(base - float(cur_total[2])) < 1) else 'grand'
        if base_mode is None:
            base_mode = 'within' if cur_total else 'grand'
        # Pass 2: 写入公式
        cur_total = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a is None or str(a).strip() in ('', '行标签'):
                continue
            label = str(a).strip()
            b = ws.cell(row=r, column=2).value
            is_num = isinstance(b, (int, float)) or (isinstance(b, str) and b.replace('.', '').replace(',', '').isdigit())
            if not is_num:
                continue
            if label in cat_set:  # 定位合计行
                cur_total = (r, label, b)
                ws.cell(row=r, column=2, value=sumif(cl[args.sales_col], f'$A{r}')).number_format = NUM
                ws.cell(row=r, column=3, value=sumif(cl[args.rev_col], f'$A{r}')).number_format = NUM2
                converted += 2
                if base_mode == 'grand':
                    ws.cell(row=r, column=4, value=f'=B{r}/SUM({s}!${cl[args.sales_col]}${L}:${cl[args.sales_col]}${R})').number_format = PCT
                    ws.cell(row=r, column=5, value=f'=C{r}/SUM({s}!${cl[args.rev_col]}${L}:${cl[args.rev_col]}${R})').number_format = PCT
                else:
                    ws.cell(row=r, column=4, value=f'=B{r}/B${r}').number_format = PCT
                    ws.cell(row=r, column=5, value=f'=C{r}/C${r}').number_format = PCT
                converted += 2
            elif ASIN_RE.match(label):  # ASIN 明细行
                ws.cell(row=r, column=2, value=sumif(cl[args.sales_col], f'$A{r}')).number_format = NUM
                ws.cell(row=r, column=3, value=sumif(cl[args.rev_col], f'$A{r}')).number_format = NUM2
                converted += 2
            else:  # 品牌明细行 (需要所属定位)
                if cur_total:
                    ws.cell(row=r, column=2, value=sumif(cl[args.sales_col], f'A{cur_total[0]}', f'$A{r}')).number_format = NUM
                    ws.cell(row=r, column=3, value=sumif(cl[args.rev_col], f'A{cur_total[0]}', f'$A{r}')).number_format = NUM2
                    converted += 2
            # 占比列 D/E (明细行)
            if label not in cat_set:
                if base_mode == 'grand':
                    ws.cell(row=r, column=4, value=f'=B{r}/SUM({s}!${cl[args.sales_col]}${L}:${cl[args.sales_col]}${R})').number_format = PCT
                    ws.cell(row=r, column=5, value=f'=C{r}/SUM({s}!${cl[args.rev_col]}${L}:${cl[args.rev_col]}${R})').number_format = PCT
                elif cur_total:
                    ws.cell(row=r, column=4, value=f'=B{r}/B${cur_total[0]}').number_format = PCT
                    ws.cell(row=r, column=5, value=f'=C{r}/C${cur_total[0]}').number_format = PCT
                converted += 2
        print(f'  [✓] {sn}: 公式化完成 (占比基数: {base_mode})')
    out = args.out or (os.path.splitext(args.input)[0] + '-公式版.xlsx')
    wb.save(out)
    print(f'[✓] saved: {out}  (写入 {converted} 个公式单元格)')
    print(f'  源表: {args.source_sheet} 数据行 {L}-{R} | 分类: {len(cat_set)} 个 | 全量: 销{grand_s:,.0f} 额{grand_r:,.0f}')


if __name__ == '__main__':
    main()
